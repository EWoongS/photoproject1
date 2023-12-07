import os
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilenames
from tkinter import messagebox
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


#사진 선택 및 경로 추출
def photoPath1():
    global remember_path

    try:
        if remember_path == '':
            filenames = askopenfilenames(initialdir="./", filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        else:
            filenames = askopenfilenames(initialdir=remember_path, filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        
        if filenames:
            for filename in filenames:
                if(filename.split(".")[-1] == "jpg" or filename.split(".")[-1] == "png"):
                    listbox5.insert(0, filename)                                                                  
                else:
                    messagebox.showerror("Error", "잘못된 양식의 파일")

            remember_path = rememberPath(filenames[0])
            
    except:
        messagebox.showerror("Error", "오류가 발생했습니다")

def photoPath2():
    global remember_path

    try:
        if remember_path == '':
            filenames = askopenfilenames(initialdir="./", filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        else:
            filenames = askopenfilenames(initialdir=remember_path, filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        
        if filenames:
            for filename in filenames:
                if(filename.split(".")[-1] == "jpg" or filename.split(".")[-1] == "png"):
                    listbox6.insert(0, filename)                                                                  
                else:
                    messagebox.showerror("Error", "잘못된 양식의 파일")

            remember_path = rememberPath(filenames[0])

    except:
        messagebox.showerror("Error", "오류가 발생했습니다")

def photoPath3():
    global remember_path

    try:
        if remember_path == '':
            filenames = askopenfilenames(initialdir="./", filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        else:
            filenames = askopenfilenames(initialdir=remember_path, filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        
        if filenames:
            for filename in filenames:
                if(filename.split(".")[-1] == "jpg" or filename.split(".")[-1] == "png"):
                    listbox7.insert(0, filename)                                                                  
                else:
                    messagebox.showerror("Error", "잘못된 양식의 파일")

            remember_path = rememberPath(filenames[0])

    except:
        messagebox.showerror("Error", "오류가 발생했습니다")

def photoPath4():
    global remember_path

    try:
        if remember_path == '':
            filenames = askopenfilenames(initialdir="./", filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        else:
            filenames = askopenfilenames(initialdir=remember_path, filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        
        if filenames:
            for filename in filenames:
                if(filename.split(".")[-1] == "jpg" or filename.split(".")[-1] == "png"):
                    listbox8.insert(0,filename)                                                                  
                else:
                    messagebox.showerror("Error", "잘못된 양식의 파일")
        
            remember_path = rememberPath(filenames[0])

    except:
        messagebox.showerror("Error", "오류가 발생했습니다")

def photoPath5():
    global remember_path

    print(label9.cget("text"))

    try:
        if remember_path == '':
            filenames = askopenfilenames(initialdir="./", filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        else:
            filenames = askopenfilenames(initialdir=remember_path, filetypes=(("jpg files", ".jpg .png "), ('All files', '*.*')))
        
        if filenames:
            for filename in filenames:
                if(filename.split(".")[-1] == "jpg" or filename.split(".")[-1] == "png"):
                    listbox9.insert(0,filename)
                else:
                    messagebox.showerror("Error", "잘못된 양식의 파일")
            
            remember_path = rememberPath(filenames[0])
            
    except:
        messagebox.showerror("Error", "오류가 발생했습니다")


def rememberPath(path):
    sep_path = path.split('/')
    temp_path = sep_path[0]
    for i in range(1, len(sep_path)-1):
        temp_path += (sep_path[i] + '/')

    return temp_path


def deleteAllPath():
    listbox5.delete(0, tk.END)
    listbox6.delete(0, tk.END)
    listbox7.delete(0, tk.END)
    listbox8.delete(0, tk.END)
    listbox9.delete(0, tk.END)

def move_save_path():
    try:
        path = os.path.realpath(txt_dest_path.get(0, tk.END)[0])
        os.startfile(path)
    except:
        messagebox.showwarning("저장 경로 미입력", "저장 경로를 지정해주세요.")

#Enrty 안에 값이 
def checkEntry():
    #각 Entry 안의 값을 확인. 없으면 경고 메세지 출력.

    if txt_dest_path.size()==0:
        messagebox.showwarning("경고", "저장 경로를 선택하세요")
        return False

    if len(entry1.get())==0:
        messagebox.showwarning("경고", "현장명을 입력하세요")
        return False

    if len(entry2.get())==0:
        messagebox.showwarning("경고", "작업명을 입력하세요")
        return False

    if len(entry3.get())==0:
        messagebox.showwarning("경고", "작업 호기를 입력하세요")
        return False

    if len(entry4.get())==0:
        messagebox.showwarning("경고", "작업일을 입력하세요")
        return False

    if (listbox5.size()+listbox6.size()+listbox7.size()+listbox8.size()) <= 0:
        messagebox.showwarning("경고", "사진을 넣어주세요")
        return False

    return True

def sumPhotoPath():
    array = []
    array.append(listbox5.get(0, listbox5.size()))
    array.append(listbox6.get(0, listbox6.size()))
    array.append(listbox7.get(0, listbox7.size()))
    array.append(listbox8.get(0, listbox8.size()))
    array.append(listbox8.get(0, listbox9.size()))

    return array



def createXL():
    #이미지 크기 3, 4, 5, 6
    img_size = [[312, 315], [312, 315], [312, 315], [312, 315]]
    img_pos = [[['A5', 'G5', 'A17'],['A6', 'G7', 'A18']],
               [['A5', 'H5', 'A19', 'H19'],['A6', 'H6', 'A20', 'H20']],
               [['A5', 'E5', 'I5', 'A16', 'G16'],['A6', 'E6', 'I6', 'A17', 'G17']],
               [['A5', 'G5', 'A14', 'G14', 'A23', 'G23'],['A6', 'G6', 'A15', 'G15', 'A24', 'G24']]]
    select_img_pos = []
    openfile_num = 0

    filetype = combo1.get()
    if filetype == '3장 사진대지':
        load_file = load_workbook("C:/AutoPhotoGenerator/protofile/프로토타입 - 3.xlsx",data_only=True)
        load_sheet = load_file.active
        openfile_num = 3
        select_img_pos = img_pos[0]
    if filetype == '4장 사진대지':
        load_file = load_workbook("C:/AutoPhotoGenerator/protofile/프로토타입 - 4.xlsx",data_only=True)
        load_sheet = load_file.active
        openfile_num = 4
        select_img_pos = img_pos[1]
    if filetype == '5장 사진대지':
        load_file = load_workbook("C:/AutoPhotoGenerator/protofile/프로토타입 - 5.xlsx",data_only=True)
        load_sheet = load_file.active
        openfile_num = 5
        select_img_pos = img_pos[2]
    if filetype == '6장 사진대지':
        load_file = load_workbook("C:/AutoPhotoGenerator/protofile/프로토타입 - 6.xlsx",data_only=True)
        load_sheet = load_file.active
        openfile_num = 6
        select_img_pos = img_pos[3]
    

    #파일 안의 이미지 및 상태 삭제-----------------------
    copy_sheet_images = load_sheet._images[:] 

    for image in copy_sheet_images:
        load_sheet._images.remove(image)

    for photo_status in select_img_pos[1]:
        load_sheet[photo_status] = ''
    #-------------------------------------------

    #각 사진파일 경로 불러오기
    all_photo_path = sumPhotoPath()
    stack_photo = 0
    stack_XL = 1

    #현장명 + 작업명
    load_sheet['A1'] = entry1.get() + " " + entry2.get()
    #작업호기
    load_sheet['A3'] = "1. 작업호기: " + entry3.get()
    #작업일
    load_sheet['A4'] = "2. 작 업 일 : " + entry4.get()

    # 여기서부터 수리-------------------------------------------------------------------
    for index in range(len(all_photo_path)):
        for photo_path in all_photo_path[index]:
            img = Image(photo_path)
            img.height = img_size[0]
            img.width = img_size[1]
            if stack_photo >= openfile_num:
                #사진이 4장이 넘어갈 경우 저장 해주고 새로운 파일 만듦
                load_file.save(txt_dest_path.get(0, txt_dest_path.size())[0] + "/" + load_sheet['A1'].value + " " +str(stack_XL)+".xlsx")

                #첫번째에 사진 넣음
                load_sheet.add_image(img, select_img_pos[1][0])
                sheet_position = select_img_pos[0][0]

                stack_photo = 1
                stack_XL += 1
            
            #이미지 넣음
            else:
                if stack_photo == 0:
                    load_sheet.add_image(img, 'A11')
                    sheet_position = 'A10'
                elif stack_photo == 1:
                    load_sheet.add_image(img, 'H11')
                    sheet_position = 'H10'
                elif stack_photo == 2:
                    load_sheet.add_image(img, 'A25')
                    sheet_position = 'A24'
                else:
                    load_sheet.add_image(img, 'H25')
                    sheet_position = 'H24'

                stack_photo += 1


            # 사진 위 상태표시
            if index == 0:
                load_sheet[sheet_position] = str(index+1) + ") 교체전"
            elif index == 1:
                load_sheet[sheet_position] = str(index+1) + ") 교체중"
            elif index == 2:
                load_sheet[sheet_position] = str(index+1) + ") 교체후"
            else:
                load_sheet[sheet_position] = str(index+1) + ") 폐기물"

    load_file.save(txt_dest_path.get(0, txt_dest_path.size())[0] + "/" + load_sheet['A1'].value + " " +str(stack_XL)+".xlsx")




def start():
    if checkEntry():
        try:
            createXL()
        except:
            messagebox.showwarning("오류", "원본 파일이 존재하지 않습니다")


# 저장 경로(폴더)
def browse_dest_path():
    forlder_selected = filedialog.askdirectory()
    if forlder_selected =='':
        return
    #print(forlder_selected)
    txt_dest_path.delete(0, tk.END)
    txt_dest_path.insert(0, forlder_selected)



#전역 변수
root = tk.Tk()
root.configure(bg='white')
remember_path = ''

#화면 초기 설정
root.title("사진대지 제작")
root.geometry("640x450+200+200")
root.resizable(False, False)

#화면 디스플레이
label1 = tk.Label(root, text="현장명     : ", bg = "white")
label1.place(x = 30, y = 50)
entry1 = tk.Entry(fg="black", bg = "snow", width = 25, background="lightgray")
entry1.place(x = 100, y = 50)

# 작업명
label2 = tk.Label(root, text="작업명     : ", bg = "white")
label2.place(x = 30, y = 90)
entry2 = tk.Entry(fg="black", bg = "snow", width = 25, background="lightgray")
entry2.place(x = 100, y = 90)

# 작업 호기
label3 = tk.Label(root, text="작업 호기 : ", bg = "white")
label3.place(x = 30, y = 130)
entry3 = tk.Entry(fg="black", bg = "snow", width = 25, background="lightgray")
entry3.place(x = 100, y = 130)

# 작업일
label4 = tk.Label(root, text="작업일     : ", bg = "white")
label4.place(x = 30, y = 170)
entry4 = tk.Entry(fg="black", bg = "snow", width = 25, background="lightgray")
entry4.place(x = 100, y = 170)

#사진 경로 스크롤바(작업전, 작업중, 작업후, 폐기물)
listbox5 = tk.Listbox(root, height=4, width=30, background="lightgray")
listbox5.place(x = 400, y = 50)
label5 = tk.Label(root, text="작업전 : ", bg = "white")
label5.place(x = 330, y = 50)
button5 = tk.Button(root, text = "사진 찾기", command=photoPath1)
button5.place(x = 330, y = 70)

listbox6 = tk.Listbox(root, height=4, width=30, background="lightgray")
listbox6.place(x = 400, y = 120)
label6 = tk.Label(root, text="작업중 : ", bg = "white")
label6.place(x = 330, y = 120)
button6 = tk.Button(root, text = "사진 찾기", command=photoPath2)
button6.place(x = 330, y = 140)

listbox7 = tk.Listbox(root, height=4, width=30, background="lightgray")
listbox7.place(x = 400, y = 190)
label7 = tk.Label(root, text="작업후 : ", bg = "white")
label7.place(x = 330, y = 190)
button7 = tk.Button(root, text = "사진 찾기", command=photoPath3)
button7.place(x = 330, y = 210)

listbox8 = tk.Listbox(root, height=4, width=30, background="lightgray")
listbox8.place(x = 400, y = 260)
label8 = tk.Label(root, text="폐기물 : ", bg = "white")
label8.place(x = 330, y = 260)
button8 = tk.Button(root, text = "사진 찾기", command=photoPath4)
button8.place(x = 330, y = 280)

listbox9 = tk.Listbox(root, height=4, width=30, background="lightgray")
listbox9.place(x = 400, y = 330)
label9 = tk.Label(root, text="입고자재 : ", bg = "white")
label9.place(x = 330, y = 330)
button9 = tk.Button(root, text = "사진 찾기", command=photoPath5)
button9.place(x = 330, y = 350)


#저장경로 설정
txt_dest_path = tk.Listbox(fg="black", bg = "snow", width = 25, background="lightgray", height=1)
txt_dest_path.insert(0, "C:/Users/Home/Desktop")
txt_dest_path.place(x = 30, y = 220)

button9 = tk.Button(root, text = "찾아보기", command=browse_dest_path, width= 9)
button9.place(x = 210, y = 215)

#저장 폴더 이동 버튼
button9 = tk.Button(root, text = "저장 폴더로 이동", command=move_save_path, width= 35)
button9.place(x = 30, y = 270)

#사진 전체 조기화
button9 = tk.Button(root, text = "사진 경로 전체 초기화", command=deleteAllPath, width= 35)
button9.place(x = 30, y = 310)

#사진 대지 생성 버튼
combo1 = ttk.Combobox(root, values=["3장 사진대지", "4장 사진대지", "5장 사진대지", "6장 사진대지"], state='readonly' ,width=14)
combo1.current(0)
combo1.place(x = 30, y = 355)

#사진 대지 생성 버튼
combo2 = ttk.Combobox(root, values=["단일호기", "복수호기"], state='readonly' ,width=14)
combo2.current(0)
combo2.place(x = 163, y = 355)

#사진 대지 생성 버튼
button9 = tk.Button(root, text = "사진대지 생성", command=start, width= 35)
button9.place(x = 30, y = 390)

root.mainloop()
